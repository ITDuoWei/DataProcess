import openpyxl

source_tem = "TeamTemplate.xlsx"

wb_tem = openpyxl.load_workbook(source_tem)
ws_tem = wb_tem.worksheets[0]

# wb_tem.create_sheet("sheet2")
# ws_tem2 = wb_tem.worksheets[1]

# print(ws_tem)
# 获取第一列
# print( t.value for t in ws_tem["A:A"])

r = 2
while r <= ws_tem.max_row:
    team = str(ws_tem.cell(r, 1).value)
    ItemDescription = ws_tem.cell(r, 2).value
    construction = "construction"
    Construction = "Construction"
    Bus = "Bus"
    Truck = "Truck"
    tradeType = ""

    if team.find(Construction) >= 0:
        classify = "Construction"
    elif team.find(construction) >= 0:
        classify = "Construction"
    elif team.find(Bus) >= 0:
        classify = "Bus"
    elif team.find(Truck) >= 0:
        classify = "Truck"

    if team.find("Domestic") >= 0:
        tradeType = "Domestic"
    elif team.find(("Export")) >= 0:
        tradeType = "Export"
    elif team.find("Taiwan") >= 0:
        tradeType = "Taiwan"
    print(team)
    print(tradeType)
    ws_tem.cell(r, 7, tradeType)

    ws_tem.cell(r, 3, team)
    ws_tem.cell(r, 4, classify)
    ws_tem.cell(r, 5, ItemDescription)

    team = team.replace("Domestic", " ")
    team = team.replace("Export", " ")
    team = team.replace("Truck", " ")
    team = team.replace("Construction", " ")
    team = team.replace("construction", " ")
    team = team.replace("Bus", " ")

    ws_tem.cell(r, 6, team.strip())

    r += 1

wb_tem.save("TeamTemplate.xlsx")

# list_CustomerType = [g.value for g in ws_new['G:G']]

# [str(col.value)[0:3] for col in ws_new['Q:Q']]
