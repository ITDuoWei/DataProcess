import os
# import sys
# print(sys.path)

# 导入openpyxl用于处理Excel
# 导入Font用于处理字体样式
# 导入reyong用于处理正则表达式
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

import time

import datetime

starttime = datetime.datetime.now()

# source/destinat
source = "source.xlsx"
destinat = "destination.xlsx"

print("Get data source successfully---" + source)

# 功能一:读取制定的工作簿
# --------------------------------------------------------------------------------

# 获取数据源
# step1 获取源工作簿的工作表  ,  data_only=True 可以忽略公式,直接提取值
wb = openpyxl.load_workbook(source, data_only=True)
# 获取Sheet表格
ws = wb.worksheets[0]

# 创建目标文件,并获取工作簿  Sheet
# --------------------------------------------------------------------------------
# step2 新建工作簿
nwb = openpyxl.Workbook()
nws = nwb.worksheets[0]
nwb.create_sheet("ESN_repeat")

ws_max_row = ws.max_row
ws_max_col = ws.max_column
print(str(ws_max_row) + " records " + str(ws_max_col) + " columns will be processed for you this time")

# 功能二:筛选数据
# (筛选“Inventory Org. ”  的值为 'BUS', 'CNS', 'CON', 'DOM', 'EXP', 'OHW','ACCRUAL','DBU' 的记录)
# --------------------------------------------------------------------------------
ls_table = list(ws.values)
ls_table2 = [t for t in ls_table if
             t[2] in ['Inventory Org.', 'BUS', 'CNS', 'CON', 'DOM', 'EXP', 'OHW', 'ACCRUAL', 'DBU']]

# step4 循环写入新的工作表
for r in ls_table2:
    nws.append(r)
nwb.save(destinat)

print("Target created successfully---" + destinat)

# 功能三:Customer Type - by No. 补位处理
# --------------------------------------------------------------------------------

# wb_new 目标工作簿
wb_new = openpyxl.load_workbook(destinat, data_only=True)
# wb_new 目标工作簿的第一个工作表
ws_new = wb_new.worksheets[0]

# 获取第六列

# G列(Customer Type - by No.)列补位操作
list_CustomerType = [g.value for g in ws_new['G:G']]

# I列(Bill-to)补位补位操作
list_Bill_to = [i.value for i in ws_new['I:I']]

# J列(Ship-to)补位补位操作
list_Ship_to = [j.value for j in ws_new['J:J']]

# Q列(Invoice Number) 补位
list_InvoiceNumber = [q.value for q in ws_new['Q:Q']]

# 设置字体样式红色加粗 用于标记Config为空的记录
font = Font(bold=True, color='FF0000')

# 取列 Invoice Number 的前三位
list_Invoice_Number_pre3 = [str(col.value)[0:3] for col in ws_new['Q:Q']]

# 新增AZ列存放RC
nrows = ws_new.max_column
# 写表头
ws_new.cell(1, nrows + 1, 'RC')
# Invoice_Date/Invoice_GL_Date时间格式化 time.strptime
ts_Invoice_Date = ""
ts_Invoice_GL_Date = ""

# 首行是标题行跳过不处理
r = 2
while r <= ws_new.max_row:
    # 写入第七列 Customer Type - by No.
    ws_new.cell(r, 7, '0' + str(list_CustomerType[r - 1]))

    # 写入第九列 Bill-to
    str_Bill_to = str(list_Bill_to[r - 1])
    if len(str_Bill_to) == 1:
        ws_new.cell(r, 9, "0000" + str_Bill_to)
    elif len(str_Bill_to) == 2:
        ws_new.cell(r, 9, "000" + str_Bill_to)
    elif len(str_Bill_to) == 3:
        ws_new.cell(r, 9, "00" + str_Bill_to)
    elif len(str_Bill_to) == 4:
        ws_new.cell(r, 9, "0" + str_Bill_to)

    # 写入第十列Ship-to
    str_Ship_to = str(list_Ship_to[r - 1])
    if len(str_Ship_to) == 1:
        ws_new.cell(r, 10, "0000" + str_Ship_to)
    elif len(str_Ship_to) == 2:
        ws_new.cell(r, 10, "000" + str_Ship_to)
    elif len(str_Ship_to) == 3:
        ws_new.cell(r, 10, "00" + str_Ship_to)
    elif len(str_Ship_to) == 4:
        ws_new.cell(r, 10, "0" + str_Ship_to)

    # 写入第十七列 Invoice Number
    str_Invoice_Number = str(list_InvoiceNumber[r - 1])
    if len(str_Invoice_Number) > 9:
        ws_new.cell(r, 17, str_Invoice_Number[0:3] + str_Invoice_Number[4:])

    #  功能九 判断ESN是否超过八位，如果超过八位则进行替换
    str_ESN = str(ws_new.cell(r, 35).value)
    if len(str_ESN) > 8:
        ws_new.cell(r, 35, re.sub('\D', '', str_ESN))

    # 功能八:列名为“ Config #- ” 为空的记录筛选出来
    if ws_new.cell(r, 30).value is None:
        col = 1
        while col <= ws_new.max_column:
            ws_new.cell(r, col).font = font
            col += 1

    # 追加一列写入RC值
    ws_new.cell(r, nrows + 1, list_Invoice_Number_pre3[r - 1])

    # GM% 列的处理，数据源是计算列公式结果是  10% 按照去公式处理 读取到 0.1 需要*100
    GM = ws_new.cell(r, 44).value * 100
    ws_new.cell(r, 44, str(GM) + "%")

    ts_Invoice_Date = time.strptime(str(ws_new.cell(r, 18).value), "%Y-%m-%d %H:%M:%S")
    ts_Invoice_GL_Date = time.strptime(str(ws_new.cell(r, 19).value), "%Y-%m-%d %H:%M:%S")

    # Invoice Date/Invoice GL Date 日期格式化 YY/m/d
    ws_new.cell(r, 18,
                str(ts_Invoice_Date.tm_year) + "/" + str(ts_Invoice_Date.tm_mon) + "/" + str(ts_Invoice_Date.tm_mday))
    ws_new.cell(r, 19, str(ts_Invoice_GL_Date.tm_year) + "/" + str(ts_Invoice_GL_Date.tm_mon) + "/" + str(
        ts_Invoice_GL_Date.tm_mday))

    r += 1

# 功能十:退货流程
# --------------------------------------------------------------------------------


# 功能十二:新增列"Team"列，功能待讨论
# --------------------------------------------------------------------------------

# 功能十三:处理完成后，把ESN重复的做标记或者筛选出来
# --------------------------------------------------------------------------------
# step1 定义空列表：list_ESN,定义空集合：set_ESN_repeat
# step2 把ESN逐个存放值列表中,
# setp3 如果存放时发现已经存在了ESN,则把重复的ESN放入集合中
# step4 定义列表 list_ESN_repeat 过滤并存放 Sheet 表中已经存在的ESN记录
# step5 将list_ESN_repeat的记录追加方式写入ws_new_esn

print("Filtering duplicate ESN records for you...")

list_ESN = ['ESN']
set_ESN_repeat = set()
r_ESN = 1
while r_ESN < ws_new.max_row:
    # 如果当前值在重复列表中不存在则追加到 list_ESN
    if (ws_new.cell(r_ESN, 35).value not in list_ESN):
        list_ESN.append(ws_new.cell(r_ESN, 35).value)
    else:
        set_ESN_repeat.add(ws_new.cell(r_ESN, 35).value)
    r_ESN += 1

# ESN重复记录
list_ESN_repeat = [t for t in ws_new.values if t[34] in set_ESN_repeat]

# 写入到ws_new_esn表中
ws_new_esn = wb_new.worksheets[1]
for r in list_ESN_repeat:
    ws_new_esn.append(r)

# 存放集合-重复ESN记录中sales_order_number最大的那个
set_ESN_repeat.discard('ESN')
set_ESN_repeat_max_sales_order_number = set()
# 定义一个空map存放ESN-Sales Order Number

print(set_ESN_repeat)

# r = 2
# for t in set_ESN_repeat:
#     print(t)
#     dict_esn_so_number[t] = ""
#     r += 1

# print(dict_esn_so_number)


# 根据set_ESN_repeat集合的找到重复记录所在行号，然后删除
# 注意循环是从最后一行至首行遍历的，因为删除记录时会使记录数变更
# step1 集合set_ESN_repeat移除'ESN'元素
# step2 遍历判断每行的ESN是否在集合set_ESN_repeat中
# step3 获取行号，删除记录


# 定义map dict_esn_so_number 存放 [ ESN:[SONumber,SONumber[-4:]] ]
# 存储最大的
list_max_so_number = []
dict_esn_so_number = {}
list_max_so_number0 = []
list_max_so_number1 = []
list_max_so_number2 = []
list_ESN_repeat.pop(0)
for t in list_ESN_repeat:
    max_so_number = 0
    # 得到SONumber后四位的最大值对应的ESN-SONumber,添加到map(dict_esn_so_number)
    if int(str(t[13])[-4:]) > max_so_number:
        max_so_number = int(str(t[13])[-4:])
        list_max_so_number0 = t[34]
        list_max_so_number1 = t[13]
        list_max_so_number2 = max_so_number
        dict_esn_so_number[list_max_so_number0] = [list_max_so_number1, list_max_so_number2]

print(dict_esn_so_number)

# 获取集合：SONumber后四位最大值对应的SONumber
for key in dict_esn_so_number:
    set_ESN_repeat_max_sales_order_number.add(str(dict_esn_so_number[key][0]))

print(set_ESN_repeat_max_sales_order_number)

# 删除的条件 1、ESN在set_ESN_repeat 2、Sales Order Number不在list_max_so_number

r_del = ws_new.max_row

while r_del >= 1:
    if ws_new.cell(r_del, 35).value in set_ESN_repeat and str(ws_new.cell(r_del,14).value) not in set_ESN_repeat_max_sales_order_number:

        # print(str(ws_new.cell(r_del, 14)) not in set_ESN_repeat_max_sales_order_number )
        # print(ws_new.cell(r_del,1).value)
        ws_new.delete_rows(r_del)
    r_del -= 1


# 保存
# --------------------------------------------------------------------------------
wb_new.save(destinat)

# input("Press <enter>")
# print("--success--")
# os.system("pause")

endtime = datetime.datetime.now()

print("Done! Use seconds " + str((endtime - starttime).seconds))
