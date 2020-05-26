import openpyxl

source_tem = "TeamTemplate.xlsx"

wb_tem = openpyxl.load_workbook(source_tem)
ws_tem = wb_tem.worksheets[0]

# 按行获取模板表
# list_template = (list(ws_tem.values))
#
# r = 2
# while r <= ws_tem.max_row:
#     print(list_template[r - 1][0])
#     print(list_template[r - 1][1])
#     print(list_template[r - 1][2])
#     print(list_template[r - 1][4])
# r += 1

print(ws_tem.cell(2,2).value)

if not ws_tem.cell(2,2).value is None:
    print("非空")
else:
    print("空")



s = "ISF3.8s5154T"

print( str.upper(s) )
