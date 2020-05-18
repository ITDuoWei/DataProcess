# 创建工作表
import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet()
wb.create_sheet()
wb.create_sheet()
wb.create_sheet("工资表",1)
wb.save("demo1.xlsx")