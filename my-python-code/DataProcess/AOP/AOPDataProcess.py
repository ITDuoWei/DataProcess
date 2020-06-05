import openpyxl

# import re

source = "source.xlsx"
wb = openpyxl.load_workbook(source, data_only=True)
ws = wb.worksheets[0]

print("Checking start...")

# 校验内容是否是从第四行开始

check_AccMgr = str(ws.cell(3, 1).value) != "AccMgr"
check_CustName = str(ws.cell(3, 2).value) != "Cust Name"
check_CustCode = str(ws.cell(3, 3).value) != "Cust Code"
check_FCG = str(ws.cell(3, 4).value) != "FCG"
check_Plant = str(ws.cell(3, 5).value) != "Plant"
check_Range = str(ws.cell(3, 6).value) != "Range"
check_Model = str(ws.cell(3, 7).value) != "Model"
check_SO = str(ws.cell(3, 8).value) != "SO"
check_Product_ID = str(ws.cell(3, 9).value) != "Product ID"
check_Config = str(ws.cell(3, 10).value) != "Config"
check_Spec = str(ws.cell(3, 11).value) != "Spec"
check_EngineFamily = str(ws.cell(3, 12).value) != "Engine Family"
check_Application = str(ws.cell(3, 13).value) != "Application"
check_Emission = str(ws.cell(3, 14).value) != "Emission"
check_stop = False

if (check_AccMgr and check_CustName and check_CustCode and check_FCG \
        and check_Plant and check_Range and check_Model and check_SO
        and check_Product_ID and check_Config and check_Spec \
        and check_EngineFamily and check_Application and check_Emission):
    print("Please make sure the fourth line starts with data")
    check_stop = True
else:
    if check_AccMgr:
        print("Please check that column A is AccMgr")
    if check_CustName:
        print("Please check that column B is Cust Name")
    if check_CustCode:
        print("Please check that column C is Cust Code")
    if check_FCG:
        print("Please check that column D is FCG")
    if check_Plant:
        print("Please check that column E is Plant")
    if check_Range:
        print("Please check that column F is Range")
    if check_Model:
        print("Please check that column G is Model")
    if check_SO:
        print("Please check that column H is SO")
    if check_Product_ID:
        print("Please check that column I is Product ID")
    if check_Config:
        print("Please check that column J is Config")
    if check_Spec:
        print("Please check that column K is Spec")
    if check_EngineFamily:
        print("Please check that column L is EngineFamily")
    if check_Application:
        print("Please check that column M is Application")
    if check_Emission:
        print("Please check that column N is Emission")

# 除了A-N列之外列数必须是13 整数倍
# 最大列号码
max_column = ws.max_column
max_row = ws.max_row
if (max_column - 14) % 13 != 0:
    print("Make sure that the number of columns after N is a multiple of 13！！！")
    check_stop = True

if check_stop:
    input("Exit and try again when you checked")


# 检查内容是否为数字或者空

# 列的索引与列名对应函数
def convertToTitle(n: int) -> str:
    return ('' if n <= 26 else convertToTitle((n - 1) // 26)) + chr((n - 1) % 26 + ord('A'))


# 定义空字典用于存放A-N列
dictID = {}

r = 3
while r <= max_row:
    # 检查A-N是否有重复记录
    if r > 3:
        row_ID = ws.iter_rows(max_row=r, min_row=r, min_col=1, max_col=14)
        ID = str( [([str(v.value).replace(" ", "") for v in row]) for row in row_ID] )
        if ID in dictID:
            print( "Repeat Line " + str(dictID[ID]) + ' and Line ' + str(r))
        else:
            dictID[ID] = r

    c = 1
    while c <= max_column:
        cell_value = ws.cell(r, c).value

        # 标题行中FY后面需要跟 YYYY 格式年份
        if r == 3 and c >= 15 and str(cell_value).upper().find("FY") >= 0 and (
                str(cell_value)[0:2] != "FY" or len(str(cell_value)[2:]) != 4):
            print("Please check row: " + str(r) +" column: " + convertToTitle(c) + " , make sure the format : FY+YYYY")
        # 检查内容是否为数值
        if r > 3 and c >= 15 and type(cell_value) != int and cell_value is not None:
            print("Please check row: " + str(r) + " column: " + convertToTitle(c) + " , The value is not a number ")

        c += 1

    r += 1

input("Checking end")