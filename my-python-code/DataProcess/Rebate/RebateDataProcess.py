import traceback

import openpyxl
import datetime

starttime = datetime.datetime.now()

print("Please close the rebateTarget.xlsx")

try:
    table_desa = "destination.xlsx"
    table_rebate = "rebateTarget.xlsx"

    # 获取工作簿
    wb_desa = openpyxl.load_workbook(table_desa)
    wb_rebate = openpyxl.load_workbook(table_rebate)

    #
    # 获取工作表
    ws_desa = wb_desa.worksheets[0]
    ws_rebate = wb_rebate.worksheets[0]

    list_desa = (list(ws_desa.values))
    list_MTBECE = []
    list_TBECE = []

    print("loading DESA data...")

    month = ws_desa['A':'A']
    list_month = [m.value for m in month]

    team = ws_desa['AT':'AT']
    list_team = [t.value for t in team]

    lob = ws_desa['AU':'AU']
    list_lob = [l.value for l in lob]

    emission = ws_desa['AV':'AV']
    list_emission = [e.value for e in emission]

    customer = ws_desa['AW':'AW']
    list_customer = [c.value for c in customer]

    enginefamily = ws_desa['AX':'AX']
    list_enginefamily = [e.value for e in enginefamily]

    list_MTBECE_zip = list(zip(list_month, list_team, list_lob, list_emission, list_customer, list_enginefamily))
    list_TBECE_zip = list(zip(list_team, list_lob, list_emission, list_customer, list_enginefamily))
    list_MTBECE = [list(MTBECE) for MTBECE in list_MTBECE_zip]
    list_TBECE = [list(TBECE) for TBECE in list_TBECE_zip]

    listJan = ["Jan"]
    listFeb = ["Feb"]
    listMar = ["Mar"]
    listApr = ["Apr"]
    listMay = ["May"]
    listJun = ["Jun"]
    listJul = ["Jul"]
    listAug = ["Aug"]
    listSep = ["Sep"]
    listOct = ["Oct"]
    listNov = ["Nov"]
    listDec = ["Dec"]
    # 获取当前的月份
    month = datetime.datetime.now().month
    # 每个月的折扣数量
    debatecount = 0

    # 判断数据部分是否为空
    rebate_data = ws_rebate['F':'Q']
    list_rebate_data = [[r.value for r in rn][1:] for rn in rebate_data]
    Noneflag = False
    if list_rebate_data == [[None, None], [None, None], [None, None], [None, None], [None, None], [None, None],
                            [None, None], [None, None], [None, None], [None, None], [None, None], [None, None]]:
        Noneflag = True

    # 循环处理rebate数据  Customer,Engine Family,Emission,LOB,Team
    list_TBECE_rebate = []
    r = 2
    while r <= ws_rebate.max_row:
        if r % 100 == 0:
            print("Passed " + str(r) + " records,surplus " + str(int(ws_rebate.max_row - (r / 100) * 100)) + " records")

        list_TBECE_rebate = [ws_rebate.cell(r, 5).value, ws_rebate.cell(r, 4).value, ws_rebate.cell(r, 3).value,
                             ws_rebate.cell(r, 1).value,
                             ws_rebate.cell(r, 2).value]

        # 统计每个月的符合条件的数量
        debatecount1 = list_MTBECE.count(listJan + list_TBECE_rebate)
        debatecount2 = list_MTBECE.count(listFeb + list_TBECE_rebate)
        debatecount3 = list_MTBECE.count(listMar + list_TBECE_rebate)
        debatecount4 = list_MTBECE.count(listApr + list_TBECE_rebate)
        debatecount5 = list_MTBECE.count(listMay + list_TBECE_rebate)
        debatecount6 = list_MTBECE.count(listJun + list_TBECE_rebate)
        debatecount7 = list_MTBECE.count(listJul + list_TBECE_rebate)
        debatecount8 = list_MTBECE.count(listAug + list_TBECE_rebate)
        debatecount9 = list_MTBECE.count(listSep + list_TBECE_rebate)
        debatecount10 = list_MTBECE.count(listOct + list_TBECE_rebate)
        debatecount11 = list_MTBECE.count(listNov + list_TBECE_rebate)
        debatecount12 = list_MTBECE.count(listDec + list_TBECE_rebate)

        # 如果数据值全部为空,则说明首次使用，直接更新所有数据
        if Noneflag:

            ws_rebate.cell(r, 6, debatecount1)
            ws_rebate.cell(r, 7, debatecount2)
            ws_rebate.cell(r, 8, debatecount3)
            ws_rebate.cell(r, 9, debatecount4)
            ws_rebate.cell(r, 10, debatecount5)
            ws_rebate.cell(r, 11, debatecount6)
            ws_rebate.cell(r, 12, debatecount7)
            ws_rebate.cell(r, 13, debatecount8)
            ws_rebate.cell(r, 14, debatecount9)
            ws_rebate.cell(r, 15, debatecount10)
            ws_rebate.cell(r, 16, debatecount11)
            ws_rebate.cell(r, 17, debatecount12)
        else:
            # 如果统计到的记录大于0则更新
            if list_TBECE.count(list_TBECE_rebate) > 0:
                if month == 1:
                    ws_rebate.cell(r, 6, debatecount1)
                if month == 2:
                    ws_rebate.cell(r, 7, debatecount2)
                if month == 3:
                    ws_rebate.cell(r, 8, debatecount3)
                if month == 4:
                    ws_rebate.cell(r, 9, debatecount4)
                if month == 5:
                    ws_rebate.cell(r, 10, debatecount5)
                if month == 6:
                    ws_rebate.cell(r, 11, debatecount6)
                if month == 7:
                    ws_rebate.cell(r, 12, debatecount7)
                if month == 8:
                    ws_rebate.cell(r, 13, debatecount8)
                if month == 9:
                    ws_rebate.cell(r, 14, debatecount9)
                if month == 10:
                    ws_rebate.cell(r, 15, debatecount10)
                if month == 11:
                    ws_rebate.cell(r, 16, debatecount11)
                if month == 12:
                    ws_rebate.cell(r, 17, debatecount12)

        # 计算季度、全年的数量
        countJan = ws_rebate.cell(r, 6).value
        countFeb = ws_rebate.cell(r, 7).value
        countMar = ws_rebate.cell(r, 8).value
        countApr = ws_rebate.cell(r, 9).value
        countMay = ws_rebate.cell(r, 10).value
        countJun = ws_rebate.cell(r, 11).value
        countJul = ws_rebate.cell(r, 12).value
        countAug = ws_rebate.cell(r, 13).value
        countSep = ws_rebate.cell(r, 14).value
        countOct = ws_rebate.cell(r, 15).value
        countNov = ws_rebate.cell(r, 16).value
        countDec = ws_rebate.cell(r, 17).value

        # 校验如果值为空设置为0

        if countJan is None:
            countJan = 0
        if countFeb is None:
            countFeb = 0
        if countMar is None:
            countMar = 0
        if countApr is None:
            countApr = 0
        if countMay is None:
            countMay = 0
        if countJun is None:
            countJun = 0
        if countJul is None:
            countJul = 0
        if countAug is None:
            countAug = 0
        if countSep is None:
            countSep = 0
        if countOct is None:
            countOct = 0
        if countNov is None:
            countNov = 0
        if countDec is None:
            countDec = 0

        try:
            countQ1 = countJan + countFeb + countMar
            countQ2 = countApr + countMay + countJun
            countQ3 = countJul + countAug + countSep
            countQ4 = countOct + countNov + countDec
            countFY = countQ1 + countQ2 + countQ3 + countQ4
            if month == 1:
                countYTD = countJan
            if month == 2:
                countYTD = countJan + countFeb
            if month == 3:
                countYTD = countQ1
            if month == 4:
                countYTD = countQ1 + countApr
            if month == 5:
                countYTD = countQ1 + countApr + countMay
            if month == 6:
                countYTD = countQ1 + countQ2
            if month == 7:
                countYTD = countQ1 + countQ2 + countJul
            if month == 8:
                countYTD = countQ1 + countQ2 + countJul + countAug
            if month == 9:
                countYTD = countQ1 + countQ2 + countQ3
            if month == 10:
                countYTD = countQ1 + countQ2 + countQ3 + countOct
            if month == 11:
                countYTD = countQ1 + countQ2 + countQ3 + countOct + countNov
            if month == 12:
                countYTD = countQ1 + countQ2 + countQ3 + countQ4

        except:
            print("Make sure the data area is numeric")

        # Q1
        ws_rebate.cell(r, 18, countQ1)
        # Q2
        ws_rebate.cell(r, 19, countQ2)
        # Q3
        ws_rebate.cell(r, 20, countQ3)
        # Q4
        ws_rebate.cell(r, 21, countQ4)
        # FY
        ws_rebate.cell(r, 22, countFY)
        # YTD
        ws_rebate.cell(r, 23, countYTD)

        r += 1

    wb_rebate.save(table_rebate)
    endtime = datetime.datetime.now()
    print("Done! Use seconds " + str((endtime - starttime).seconds))
    print("Mission accomplished!Please exit")
    input()
except:
    print("Mission error!Please check")
    print(traceback.format_exc())
    input()
