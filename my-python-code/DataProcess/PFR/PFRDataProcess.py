import traceback

import openpyxl
import datetime

try:

    starttime = datetime.datetime.now()

    doc = """
    
      _____       _   _                   _____                   _             
     |  __ \     | | | |                 |  __ \                 (_)            
     | |__) |   _| |_| |__   ___  _ __   | |__) |   _ _ __  _ __  _ _ __   __ _ 
     |  ___/ | | | __| '_ \ / _ \| '_ \  |  _  / | | | '_ \| '_ \| | '_ \ / _` |
     | |   | |_| | |_| | | | (_) | | | | | | \ \ |_| | | | | | | | | | | | (_| |
     |_|    \__, |\__|_| |_|\___/|_| |_| |_|  \_\__,_|_| |_|_| |_|_|_| |_|\__, |
             __/ |                                                         __/ |
            |___/                                                         |___/
    --------------Please have a cup of coffee and wait patiently--------------- 
    """
    print(doc)

    PFR_source = "PFRsource.xlsx"
    PFR_target = "PFRTarget.xlsx"
    rebate_Amount = "rebateAcount.xlsx"
    template = "TeamTemplate.xlsx"

    wb_source = openpyxl.load_workbook(PFR_source)
    ws_source = wb_source.worksheets[0]

    # 目标表
    wb_target = openpyxl.Workbook()
    ws_target = wb_target.worksheets[0]

    ls_table_source = list(ws_source.values)
    ls_table_values_source = [t for t in ls_table_source]

    for r in ls_table_source:
        ws_target.append(r)
    wb_target.save(PFR_target)

    #  PFR模板表
    wb_template = openpyxl.load_workbook(template)
    # Team  (EA EBU)
    ws_template = wb_template.worksheets[5]
    # Exchange Rate
    ws_template_rate = wb_template.worksheets[6]
    # FCG name 也就是 Customer
    ws_template_FCGName = wb_template.worksheets[7]
    # LOB
    ws_template_LOB = wb_template.worksheets[1]
    # PFR-Other Sales Adj.
    ws_template_Other_Sales_Adj = wb_template.worksheets[8]
    ws_template_Other_GM_Adj = wb_template.worksheets[9]
    ws_template_Other_SAR_Adj = wb_template.worksheets[10]

    # desa模板表
    ws_template_desa_emission = wb_template.worksheets[2]

    # rebateAccount 表
    wb_rebateAccount = openpyxl.load_workbook(rebate_Amount)
    # sales
    ws_rebateAccount_sales = wb_rebateAccount.worksheets[0]
    # Purchase
    ws_rebateAccount_purchase = wb_rebateAccount.worksheets[1]

    # 将LOB的存储在字典中，然后PFR每一行匹配字典
    dict_lob = {}
    dict_lob_turn = {}
    l = 2
    while l <= ws_template_LOB.max_row:
        ws_template_LOB_Team = str(ws_template_LOB.cell(l, 1).value)
        ws_template_LOB_LOB = str(ws_template_LOB.cell(l, 2).value)
        dict_lob[ws_template_LOB_Team] = ws_template_LOB_LOB
        l += 1

    dict_lob_turn = {v: k for k, v in dict_lob.items()}
    # print(dict_lob_turn)

    r = 2
    while r <= ws_target.max_row:
        # 输出进度
        if r % 100 == 0:
            print("Passed " + str(r) + " records,surplus " + str(int(ws_target.max_row - (r / 100) * 100)) + " records")

        ws_source_MFGPlant = str(ws_source.cell(r, 5).value).upper()
        PFR_EngineFamily = str(ws_source.cell(r, 7).value)
        ws_source_EngineFamily = PFR_EngineFamily.upper()
        ws_source_Application = str(ws_source.cell(r, 9).value).upper()
        ws_source_MBUName = str(ws_source.cell(r, 11).value).upper()
        PFR_FCGName = str(ws_source.cell(r, 16).value)
        ws_source_FCGName = PFR_FCGName.upper()

        ws_source_Units = ws_source.cell(r, 21).value
        ws_source_NetSales = ws_source.cell(r, 24).value
        ws_source_ProdCost = ws_source.cell(r, 25).value
        ws_source_Material = ws_source.cell(r, 30).value
        ws_source_Conversion = ws_source.cell(r, 31).value
        ws_source_AcctgPeriod = str(ws_source.cell(r, 4).value)

        # 1 更新 UnitPrice
        try:
            UnitPrice = ws_source_NetSales / ws_source_Units
            ws_target.cell(r, 38, UnitPrice)
        except:
            ws_target.cell(r, 38, "")

        # 2 更新Unit Cost
        # 内销
        if "BHO/CQP/DFM/XCE".find(ws_source_MFGPlant) >= 0:
            try:
                UnitCost = ws_source_Material / ws_source_Units
                ws_target.cell(r, 39, UnitCost)
            except:
                ws_target.cell(r, 39, "")
        else:  # 进口
            try:
                UnitCost = ws_source_ProdCost / ws_source_Units
                ws_target.cell(r, 39, UnitCost)
            except:
                ws_target.cell(r, 39, "")

        # 3 更新Unit GM
        try:
            UnitGM = UnitPrice - UnitCost
            ws_target.cell(r, 40, UnitGM)
        except:
            ws_target.cell(r, 40, "")

        # 4 更新UnitGM%
        try:
            UnitGM_precent = UnitGM / UnitPrice * 100
            ws_target.cell(r, 41, UnitGM_precent)
        except:
            ws_target.cell(r, 41, "")

        # 5 更新RC
        # RC #  ws_source_Application = “CONSTRUCTION” 为 587 其他都是 497
        if ws_source_Application == "CONSTRUCTION":
            ws_target.cell(r, 42, 587)
        else:
            ws_target.cell(r, 42, 497)

        # 6 更新 EA EBU
        # 6.1 mapping EA EBU
        t = 2
        while t <= ws_template.max_row:
            ws_template_Category = str(ws_template.cell(t, 1).value).upper()
            ws_template_Family = str(ws_template.cell(t, 4).value).upper()
            ws_template_Application = str(ws_template.cell(t, 2).value).upper()
            ws_template_MBU = str(ws_template.cell(t, 3).value).upper()
            ws_template_EAEBU = str(ws_template.cell(t, 5).value).upper()  # Team

            # other 处理
            if ws_template_Category.find(ws_source_MFGPlant) < 0:
                ws_source_MFGPlant = "OTHER"

            if ws_template_Category.find(ws_source_MFGPlant) >= 0 \
                    and ws_source_Application == ws_template_Application \
                    and ws_source_MBUName == ws_template_MBU \
                    and (ws_source_EngineFamily == ws_template_Family or ws_template_Family is None):
                # print( ws_target.cell(r,37) )
                ws_target.cell(r, 37, ws_template_EAEBU)

            t += 1

        # 6.2 隆工特殊处理
        #  隆工特殊处理 Category 属于 BHO/CQP/DFM/XCE  and Application == CONSTRUCTION and Family
        #  隆工不需要维护
        if "BHO/CQP/DFM/XCE".find(ws_source_MFGPlant) >= 0 \
                and ws_source_Application == "CONSTRUCTION" \
                and ws_source_EngineFamily == "B6.7" \
                and ws_source_FCGName == "LONKING SHANGHAI":
            ws_target.cell(r, 37, "Domestic DCEC construction")

        # 7 更新汇率
        # 通过模板表ws_template_rate取出汇率,并更新PFR
        er = 2
        while er <= ws_template_rate.max_row:
            tem_month = str(ws_template_rate.cell(er, 1).value)
            rate = str(ws_template_rate.cell(er, 2).value)
            if (ws_source_AcctgPeriod == tem_month):
                ws_target.cell(r, 43, rate)
            er += 1

        # 8 更新 discount / Purchasing rebate
        # 8.1 获取PFR的 Acctg Period \ Engine Family \ Configuration \ FCG Name
        PFR_Month = ws_source_AcctgPeriod[-2:]
        ws_source_Configuration = str(ws_source.cell(r, 10).value)
        PFR_Emission = ""
        PFR_Customer = ""

        # 8.2 根据 TeamTemplate.xlsx 表获取 Configuration 对应的 PFR_Emission
        tt = 2
        while tt <= ws_template_desa_emission.max_row:
            # 获取 模板表 Config 对应的 Emission
            ws_template_desa_emission_config = str(ws_template_desa_emission.cell(tt, 1).value)
            ws_template_desa_emission_emission = str(ws_template_desa_emission.cell(tt, 2).value)
            # 如果 PFR 的该记录的Configuration == 模板表的 Config , 则获取Emission
            if ws_source_Configuration == ws_template_desa_emission_config:
                PFR_Emission = ws_template_desa_emission_emission
            tt += 1

        # 8.3 根据PFR模板表获取 和 FCG name 对应的 PFR_Customer
        fcg = 2
        while fcg <= ws_template_FCGName.max_row:
            # 模板表 FCG Name 对应的Customer
            ws_template_FCGName_FCGName = str(ws_template_FCGName.cell(fcg, 1).value)
            ws_template_FCGName_Customer = str(ws_template_FCGName.cell(fcg, 2).value)
            if PFR_FCGName == ws_template_FCGName_FCGName:
                PFR_Customer = ws_template_FCGName_Customer
            fcg += 1

        # 8.4 更新 discount
        # 根据月份获取 discount


        sa = 2
        while sa <= ws_rebateAccount_sales.max_row:
            # 获取 rebate表的 Customer 、 Engine Family 、 Emission
            rebate_Customer = str(ws_rebateAccount_sales.cell(sa, 1).value)
            rebate_EngineFamily = str(ws_rebateAccount_sales.cell(sa, 2).value)
            rebate_Emission = str(ws_rebateAccount_sales.cell(sa, 3).value)
            if (
                    PFR_Customer == rebate_Customer and PFR_EngineFamily == rebate_EngineFamily and PFR_Emission == rebate_Emission):
                if (PFR_Month == "01"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 6).value))
                if (PFR_Month == "02"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 7).value))
                if (PFR_Month == "03"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 8).value))
                if (PFR_Month == "04"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 9).value))
                if (PFR_Month == "05"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 10).value))
                if (PFR_Month == "06"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 11).value))
                if (PFR_Month == "07"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 12).value))
                if (PFR_Month == "08"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 13).value))
                if (PFR_Month == "09"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 14).value))
                if (PFR_Month == "10"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 15).value))
                if (PFR_Month == "11"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 16).value))
                if (PFR_Month == "12"):
                    ws_target.cell(r, 23, str(ws_rebateAccount_sales.cell(sa, 17).value))

            sa += 1

        # 9 更新 Purchasing rebate
        p = 2
        while p <= ws_rebateAccount_purchase.max_row:
            rebate_purchase_Customer = str(ws_rebateAccount_purchase.cell(p, 1).value)
            rebate_purchase_EngineFamily = str(ws_rebateAccount_purchase.cell(p, 2).value)
            rebate_purchase_Emission = str(ws_rebateAccount_purchase.cell(p, 3).value)
            if (
                    PFR_Customer == rebate_purchase_Customer and PFR_EngineFamily == rebate_purchase_EngineFamily and PFR_Emission == rebate_purchase_Emission):
                if (PFR_Month == "01"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 6).value))
                if (PFR_Month == "02"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 7).value))
                if (PFR_Month == "03"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 8).value))
                if (PFR_Month == "04"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 9).value))
                if (PFR_Month == "05"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 10).value))
                if (PFR_Month == "06"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 11).value))
                if (PFR_Month == "07"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 12).value))
                if (PFR_Month == "08"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 13).value))
                if (PFR_Month == "09"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 14).value))
                if (PFR_Month == "10"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 15).value))
                if (PFR_Month == "11"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 16).value))
                if (PFR_Month == "12"):
                    ws_target.cell(r, 45, str(ws_rebateAccount_purchase.cell(p, 17).value))

            p += 1



        # 10 通过mapping方式更新lob
        ws_source_team = str(ws_source.cell(r, 37).value)
        if ws_source_team in dict_lob:
            ws_target.cell(r, 44, dict_lob[ws_source_team])

        r += 1

    # print(dict_PFR_LOB_Month_Units)

    wb_target.save(PFR_target)

    # 11 更新 Other Sales Adj. / Other GM Adj. / SAR ， 根据LOB汇总数据对 Units 进行分摊
    # 定义列表存储每个 LOB 每月 对应的 Units 累加
    dict_PFR_LOB_Month_Units = {}
    dict_PFR_LOB_Month_Units_key = ()

    # 获取LOB
    target_lob = ws_target['AR':'AR']
    target_lob_list = [str(e.value) for e in target_lob]

    # 获取Month
    target_month = ws_target['D':'D']
    target_month_list = [str(e.value)[-2:] for e in target_month]

    # 获取units的值
    target_units = ws_target['U':'U']
    target_units_list = [str(e.value) for e in target_units]

    # list_target_lob_month_units = list(zip(target_lob_list, target_month_list, target_units_list))
    # 定义字典 dict_target_lob_month_units 以 list(LOB,Month)作为Key，以Units作为Value
    dict_target_lob_month_units = dict(zip(zip(target_lob_list, target_month_list), target_units_list))

    # 遍历字典赋值为0
    for key in dict_target_lob_month_units:
        dict_target_lob_month_units[key] = 0

    # print(list_target_lob_month_units)
    # print(dict_target_lob_month_units)

    dict_tr = {}
    # 12.1 遍历求得 每个 LOB 在各个月份的 汇总值 ， 如 LCV 1月份 汇总 2月份汇总。。。
    tr = 2
    while tr <= ws_target.max_row:
        # 获取LOB
        tr_lob = ws_target.cell(tr, 44).value
        # 获取Month
        tr_month = str(ws_target.cell(tr, 4).value)[-2:]
        # 获取units的值
        tr_units = str(ws_target.cell(tr, 21).value)

        dict_tr = {(tr_lob, tr_month): tr_units}
        # dict_target_lob_month_units  units 的总和
        # 判断是否存在 LOB-Month 为键，如果存在则累加
        if (tr_lob, tr_month) in dict_target_lob_month_units:
            dict_target_lob_month_units[(tr_lob, tr_month)] = dict_target_lob_month_units[(tr_lob, tr_month)] + int(
                tr_units)

        tr += 1

    # print("-------dict_target_lob_month_units-------")
    # print(dict_target_lob_month_units)

    # 12.2 从模板表取出 Other Sales Adj
    osa = 2
    dict_osa_jan = {}
    dict_osa_feb = {}
    dict_osa_mar = {}
    dict_osa_apr = {}
    dict_osa_may = {}
    dict_osa_jun = {}
    dict_osa_jul = {}
    dict_osa_aug = {}
    dict_osa_sep = {}
    dict_osa_oct = {}
    dict_osa_nov = {}
    dict_osa_dec = {}
    while osa <= ws_template_Other_Sales_Adj.max_row:
        osa_lob = str(ws_template_Other_Sales_Adj.cell(osa, 1).value)
        osa_sales_jan = str(ws_template_Other_Sales_Adj.cell(osa, 2).value)
        osa_sales_feb = str(ws_template_Other_Sales_Adj.cell(osa, 3).value)
        osa_sales_mar = str(ws_template_Other_Sales_Adj.cell(osa, 4).value)
        osa_sales_apr = str(ws_template_Other_Sales_Adj.cell(osa, 5).value)
        osa_sales_may = str(ws_template_Other_Sales_Adj.cell(osa, 6).value)
        osa_sales_jun = str(ws_template_Other_Sales_Adj.cell(osa, 7).value)
        osa_sales_jul = str(ws_template_Other_Sales_Adj.cell(osa, 8).value)
        osa_sales_aug = str(ws_template_Other_Sales_Adj.cell(osa, 9).value)
        osa_sales_sep = str(ws_template_Other_Sales_Adj.cell(osa, 10).value)
        osa_sales_oct = str(ws_template_Other_Sales_Adj.cell(osa, 11).value)
        osa_sales_nov = str(ws_template_Other_Sales_Adj.cell(osa, 12).value)
        osa_sales_dec = str(ws_template_Other_Sales_Adj.cell(osa, 13).value)

        # 拼出Key
        dict_osa_jan[(osa_lob, "01")] = osa_sales_jan
        dict_osa_feb[(osa_lob, "02")] = osa_sales_feb
        dict_osa_mar[(osa_lob, "03")] = osa_sales_mar
        dict_osa_apr[(osa_lob, "04")] = osa_sales_apr
        dict_osa_may[(osa_lob, "05")] = osa_sales_may
        dict_osa_jun[(osa_lob, "06")] = osa_sales_jun
        dict_osa_jul[(osa_lob, "07")] = osa_sales_jul
        dict_osa_aug[(osa_lob, "08")] = osa_sales_aug
        dict_osa_sep[(osa_lob, "09")] = osa_sales_sep
        dict_osa_oct[(osa_lob, "10")] = osa_sales_oct
        dict_osa_nov[(osa_lob, "11")] = osa_sales_nov
        dict_osa_dec[(osa_lob, "12")] = osa_sales_dec

        # 获得 Other Sales Adj. 每个LOB的每个月份的值 用于后续分摊
        dict_osa = {**dict_osa_jan, **dict_osa_feb, **dict_osa_mar, **dict_osa_apr, **dict_osa_may, **dict_osa_jun,
                    **dict_osa_jul, **dict_osa_aug, **dict_osa_sep, **dict_osa_oct, **dict_osa_nov, **dict_osa_dec}

        osa += 1

    # 12.3 从模板表取出 Other GM Adj
    oga = 2
    dict_oga_jan = {}
    dict_oga_feb = {}
    dict_oga_mar = {}
    dict_oga_apr = {}
    dict_oga_may = {}
    dict_oga_jun = {}
    dict_oga_jul = {}
    dict_oga_aug = {}
    dict_oga_sep = {}
    dict_oga_oct = {}
    dict_oga_nov = {}
    dict_oga_dec = {}
    while oga <= ws_template_Other_GM_Adj.max_row:
        oga_lob = str(ws_template_Other_GM_Adj.cell(oga, 1).value)
        oga_GM_jan = str(ws_template_Other_GM_Adj.cell(oga, 2).value)
        oga_GM_feb = str(ws_template_Other_GM_Adj.cell(oga, 3).value)
        oga_GM_mar = str(ws_template_Other_GM_Adj.cell(oga, 4).value)
        oga_GM_apr = str(ws_template_Other_GM_Adj.cell(oga, 5).value)
        oga_GM_may = str(ws_template_Other_GM_Adj.cell(oga, 6).value)
        oga_GM_jun = str(ws_template_Other_GM_Adj.cell(oga, 7).value)
        oga_GM_jul = str(ws_template_Other_GM_Adj.cell(oga, 8).value)
        oga_GM_aug = str(ws_template_Other_GM_Adj.cell(oga, 9).value)
        oga_GM_sep = str(ws_template_Other_GM_Adj.cell(oga, 10).value)
        oga_GM_oct = str(ws_template_Other_GM_Adj.cell(oga, 11).value)
        oga_GM_nov = str(ws_template_Other_GM_Adj.cell(oga, 12).value)
        oga_GM_dec = str(ws_template_Other_GM_Adj.cell(oga, 13).value)

        dict_oga_jan[(oga_lob, "01")] = oga_GM_jan
        dict_oga_feb[(oga_lob, "02")] = oga_GM_feb
        dict_oga_mar[(oga_lob, "03")] = oga_GM_mar
        dict_oga_apr[(oga_lob, "04")] = oga_GM_apr
        dict_oga_may[(oga_lob, "05")] = oga_GM_may
        dict_oga_jun[(oga_lob, "06")] = oga_GM_jun
        dict_oga_jul[(oga_lob, "07")] = oga_GM_jul
        dict_oga_aug[(oga_lob, "08")] = oga_GM_aug
        dict_oga_sep[(oga_lob, "09")] = oga_GM_sep
        dict_oga_oct[(oga_lob, "10")] = oga_GM_oct
        dict_oga_nov[(oga_lob, "11")] = oga_GM_nov
        dict_oga_dec[(oga_lob, "12")] = oga_GM_dec

        # 获得 Other GM Adj. 每个LOB的每个月份的值 用于后续分摊
        dict_oga = {**dict_oga_jan, **dict_oga_feb, **dict_oga_mar, **dict_oga_apr, **dict_oga_may, **dict_oga_jun,
                    **dict_oga_jul, **dict_oga_aug, **dict_oga_sep, **dict_oga_oct, **dict_oga_nov, **dict_oga_dec}

        oga += 1

    # 12.4 从模板表取出 SAR adjust
    sar = 2
    dict_sar_jan = {}
    dict_sar_feb = {}
    dict_sar_mar = {}
    dict_sar_apr = {}
    dict_sar_may = {}
    dict_sar_jun = {}
    dict_sar_jul = {}
    dict_sar_aug = {}
    dict_sar_sep = {}
    dict_sar_oct = {}
    dict_sar_nov = {}
    dict_sar_dec = {}
    while sar <= ws_template_Other_SAR_Adj.max_row:
        sar_lob = str(ws_template_Other_SAR_Adj.cell(sar, 1).value)
        sar_jan = str(ws_template_Other_SAR_Adj.cell(sar, 2).value)
        sar_feb = str(ws_template_Other_SAR_Adj.cell(sar, 3).value)
        sar_mar = str(ws_template_Other_SAR_Adj.cell(sar, 4).value)
        sar_apr = str(ws_template_Other_SAR_Adj.cell(sar, 5).value)
        sar_may = str(ws_template_Other_SAR_Adj.cell(sar, 6).value)
        sar_jun = str(ws_template_Other_SAR_Adj.cell(sar, 7).value)
        sar_jul = str(ws_template_Other_SAR_Adj.cell(sar, 8).value)
        sar_aug = str(ws_template_Other_SAR_Adj.cell(sar, 9).value)
        sar_sep = str(ws_template_Other_SAR_Adj.cell(sar, 10).value)
        sar_oct = str(ws_template_Other_SAR_Adj.cell(sar, 11).value)
        sar_nov = str(ws_template_Other_SAR_Adj.cell(sar, 12).value)
        sar_dec = str(ws_template_Other_SAR_Adj.cell(sar, 13).value)

        dict_sar_jan[(sar_lob, "01")] = sar_jan
        dict_sar_feb[(sar_lob, "02")] = sar_feb
        dict_sar_mar[(sar_lob, "03")] = sar_mar
        dict_sar_apr[(sar_lob, "04")] = sar_apr
        dict_sar_may[(sar_lob, "05")] = sar_may
        dict_sar_jun[(sar_lob, "06")] = sar_jun
        dict_sar_jul[(sar_lob, "07")] = sar_jul
        dict_sar_aug[(sar_lob, "08")] = sar_aug
        dict_sar_sep[(sar_lob, "09")] = sar_sep
        dict_sar_oct[(sar_lob, "10")] = sar_oct
        dict_sar_nov[(sar_lob, "11")] = sar_nov
        dict_sar_dec[(sar_lob, "12")] = sar_dec

        # 获得 Other GM Adj. 每个LOB的每个月份的值 用于后续分摊
        dict_sar = {**dict_sar_jan, **dict_sar_feb, **dict_sar_mar, **dict_sar_apr, **dict_sar_may, **dict_sar_jun,
                    **dict_sar_jul, **dict_sar_aug, **dict_sar_sep, **dict_sar_oct, **dict_sar_nov, **dict_sar_dec}

        sar += 1


    # print("-------dict_osa-------")
    # print(dict_osa)
    # print(dict_oga)
    # print(dict_sar)

    # 12.5 逐行分摊 Other Sales Adj. / Other GM Adj. / SAR
    osa_w = 2
    while osa_w <= ws_target.max_row:
        # 获取LOB
        w_lob = ws_target.cell(osa_w, 44).value
        # 获取Month
        w_month = str(ws_target.cell(osa_w, 4).value)[-2:]
        # 获取units的值
        w_units = str(ws_target.cell(osa_w, 21).value)

        dict_tr = {(w_lob, w_month): w_units}
        if (w_lob, w_month) in dict_target_lob_month_units:
            if ( dict_tr[(w_lob, w_month)] == 'None' or dict_target_lob_month_units[(w_lob, w_month)] == 'None' or dict_osa[(w_lob, w_month)] == 'None' ):
                w_osa =0
            else:
                w_osa = int(dict_tr[(w_lob, w_month)]) / int(dict_target_lob_month_units[(w_lob, w_month)]) * int(dict_osa[(w_lob, w_month)])
            ws_target.cell(osa_w, 46, w_osa)

        if (w_lob, w_month) in dict_target_lob_month_units:
            if (dict_tr[(w_lob, w_month)] == 'None' or dict_target_lob_month_units[(w_lob, w_month)] == 'None' or
                    dict_oga[(w_lob, w_month)] == 'None'):
                w_osa = 0
            else:
                w_osa = int(dict_tr[(w_lob, w_month)]) / int(dict_target_lob_month_units[(w_lob, w_month)]) * int(
                    dict_oga[(w_lob, w_month)])
            ws_target.cell(osa_w, 47, w_osa)

        if (w_lob, w_month) in dict_target_lob_month_units:
            if (dict_tr[(w_lob, w_month)] == 'None' or dict_target_lob_month_units[(w_lob, w_month)] == 'None' or
                    dict_sar[(w_lob, w_month)] == 'None'):
                w_sar = 0
            else:
                w_sar = int(dict_tr[(w_lob, w_month)]) / int(dict_target_lob_month_units[(w_lob, w_month)]) * int(
                    dict_sar[(w_lob, w_month)])
            ws_target.cell(osa_w, 48, w_sar)

        osa_w += 1

    wb_target.save(PFR_target)

    endtime = datetime.datetime.now()
    print("Done! Use seconds " + str((endtime - starttime).seconds))
    print("Mission accomplished!Please exit")


except Exception as e:
    print("Mission error!Please check")
    print(traceback.format_exc())

input()
